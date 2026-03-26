#!/usr/bin/env python3
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

HEADER_MAP = {
    "description": {"tên hàng", "ten hang", "nội dung", "noi dung", "mô tả", "mo ta", "description", "item", "product", "model"},
    "origin": {"xuất xứ", "xuat xu", "origin"},
    "unit": {"đvt", "dvt", "đơn vị", "don vi", "unit"},
    "quantity": {"sl", "số lượng", "so luong", "quantity", "qty"},
    "costPrice": {"giá nhập", "gia nhap", "đơn giá nhập", "don gia nhap", "giá vốn", "gia von", "cost price", "unit price", "đơn giá", "don gia"},
}
IGNORE_ROW_PATTERNS = [
    "cộng tiền hàng", "vat", "tổng cộng", "bằng chữ", "điều khoản", "hiệu lực báo giá", "thanh toán", "giao hàng", "bảo hành", "kính gửi", "trân trọng"
]


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def simplify(text):
    text = normalize_text(text).lower()
    repl = str.maketrans({"đ": "d", "Đ": "d"})
    return text.translate(repl)


def parse_number(value):
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace("%", "")
    text = re.sub(r"[\s,.]", "", text)
    if not text:
        return None
    try:
        num = float(text)
    except ValueError:
        return None
    return int(num) if num.is_integer() else num


def is_ignored_row(values):
    joined = simplify(" ".join(v for v in values if v))
    return any(pattern in joined for pattern in IGNORE_ROW_PATTERNS)


def detect_header(row_values):
    mapping = {}
    hits = 0
    for idx, val in enumerate(row_values):
        cell = simplify(val)
        for canonical, aliases in HEADER_MAP.items():
            if cell in aliases:
                mapping[canonical] = idx
                hits += 1
                break
    return mapping if hits >= 2 and "description" in mapping else None


def extract_items(path):
    wb = load_workbook(filename=path, data_only=True, read_only=True)
    best = {"sheet": None, "header_row": None, "items": [], "warnings": []}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        header_map = None
        header_row_idx = None
        items = []
        for r_idx, row in enumerate(rows, start=1):
            values = [normalize_text(v) for v in row]
            if not any(values):
                continue
            if header_map is None:
                maybe = detect_header(values)
                if maybe:
                    header_map = maybe
                    header_row_idx = r_idx
                continue
            if is_ignored_row(values):
                continue
            desc = values[header_map.get("description", -1)] if "description" in header_map else ""
            if not desc:
                continue
            item = {
                "description": desc,
                "origin": values[header_map.get("origin", -1)] if "origin" in header_map else "",
                "unit": values[header_map.get("unit", -1)] if "unit" in header_map else "",
                "quantity": parse_number(values[header_map.get("quantity", -1)]) if "quantity" in header_map else None,
                "costPrice": parse_number(values[header_map.get("costPrice", -1)]) if "costPrice" in header_map else None,
            }
            if item["quantity"] is None and item["costPrice"] is None:
                continue
            items.append(item)
        if len(items) > len(best["items"]):
            best = {
                "sheet": ws.title,
                "header_row": header_row_idx,
                "items": items,
                "warnings": [] if items else ["Không tìm thấy dòng hàng hợp lệ trong sheet này."],
            }
    if not best["items"]:
        best["warnings"].append("Không trích xuất được item nào từ file Excel.")
    return best


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 scripts/extract_items_from_xlsx.py '<path-to-xlsx>'", file=sys.stderr)
        sys.exit(2)
    path = Path(sys.argv[1]).expanduser().resolve()
    if not path.exists():
        print(f"File không tồn tại: {path}", file=sys.stderr)
        sys.exit(1)
    result = extract_items(path)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
